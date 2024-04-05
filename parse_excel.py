'''Программа должна из всех файлов (екселя и ворда) находящихся в указанных пользователем папкам собрать данные и
записать их в ексель таблицу (пример во вложении.... дополнительно к столбцам таблицы добавить данные о пути и имени
файла из которых взяты данные'''


import re
import openpyxl
import glob
import os
import write_in_file
import read_files

class ExcelParsing:
    data = {'фамилия': None, 'имя': None, 'отчество': None,'снилс': None,
            'серия': None, 'номер': None, 'кем выдан': None, 'дата выдачи': None,'код подразделения': None,'индекс': None,
            'телефон': None,'адрес прописки': None, 'кадастровый номер работ': None, 'номер договора': None,
            'вид работ': None,'дата договора': None,
            'стоимость работ ООО': None, 'стоимость работ ИП': None, 'файл,из которого взяты данные': None}

    def __init__(self,file):
        self.files = read_files.Open_files(file).add_files(file)['excel_files']
        if not self.files:
            self.files = [file]
    def main_call(self):
        list_dicts = []
        print(f'список всех файлов:\n{self.files}')

        for file in self.files:
            print(f'текущий файл - {file}')
            self.text_data = self.parse_excel_file_data(file)
            self.data['файл,из которого взяты данные'] = file
            self.text_data_text = ' '.join(self.text_data)
            self.fio()
            self.snils()
            self.seria_passport()
            self.number_passport()
            self.who_give()
            self.date_given()
            self.phone_number()
            self.department_code()
            self.registered_address()
            self.kad_num_work()
            self.number_passport()
            self.number_passport()
            self.cost_contract_ooo()
            self.cost_contract_ip()
            self.date_contract()
            self.working_type()
            self.contract_number()
            self.index_parse()
            list_dicts.append(self.data.copy())
            self.delete_data()
        return list_dicts
    def delete_data(self):
        for i in self.data:
            self.data[i] = None

    def parse_excel_file_data(self,file_path):
        try:
            lst = []
            # Открываем Excel файл
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[wb.sheetnames[0]]  # Получаем лист с исходными данными

            # Выводим каждую ячейку на экран
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        lst.append(str(cell))
            return lst
        except Exception as e:
            print(e)
            print("Ошибка при парсинге Excel файла:", e)
            raise PermissionError('Закройте файлы, находяищееся в папке')

    def file_or_files(self,file):
        if os.path.isfile(file):
            return [file]
        elif os.path.isdir(file):
            xlsx_files = glob.glob(os.path.join(file, '*.xlsm'))
            xlsx_files1 = glob.glob(os.path.join(file, '*.xlsx'))
            xlsx_files2 = glob.glob(os.path.join(file, '*.xls'))

            return [file for file in xlsx_files]
        else:
            print('путь к файлу указан неверно')
            raise FileNotFoundError('Ссылка на файл не найдена')


    def fio(self):
        '''Находим фамилию имя отчество заказчика'''
        shablons = r'[А-Я][а-я]+\s*[А-Я][а-я]+\s*[А-Я][а-я]+'
        fio_data = re.findall(shablons,self.text_data_text)
        if fio_data:
            lst_fio_data = fio_data[0].split(' ')
            for i in lst_fio_data:
                if i.isspace() or i == '':
                    lst_fio_data.remove(i)
            self.data['фамилия'] = lst_fio_data[0]
            self.data['имя'] = lst_fio_data[1]
            self.data['отчество'] = lst_fio_data[2]



    def snils(self):
        print('текст - ')
        name_column = self.search_column_using_re(r'\s*снилс\s*')
        print('имя колонки - ',name_column)
        text = self.check_text_search(name_column)
        print(type(text))
        shablons = r'\d{3}-\d{3}-\d{3}\s*[- ]\s*\d{2}'
        res = re.findall(shablons,text,re.I|re.DOTALL)
        if res:
            self.data['снилс'] = res[0]

    def seria_passport(self):
        text = self.check_text_search("серия")
        shablons = r' \d{2} \d{2} '
        res = re.findall(shablons, text, re.I | re.DOTALL)
        if res:
            self.data['серия'] = res[0]

    def number_passport(self):
        text = self.check_text_search("номер")
        shablons = r' \d{6}\d* '
        res = re.findall(shablons, text, re.I | re.DOTALL)
        if res:
            self.data['номер'] = res[0]

    def who_give(self):
        name_column = self.search_column_using_re(r'кем\s*|выдан\s*|кем выдан\s*')
        text = self.check_list_search(name_column)
        lst_res = []
        for i in text:
            res = re.findall(r'.{19}.*', i)
            if len(res) == 0:
                continue
            lst_res.extend(res)
        if lst_res:
            self.data['кем выдан'] = lst_res[0]


    def date_given(self):
        text = self.check_list_search('дата выдачи')
        lst_res = []
        for i in text:
            res = re.findall(r'\d+-\d+-\d+.*', i)
            if len(res) == 0:
                continue
            lst_res.append(i)
        if lst_res:
            self.data['дата выдачи'] = lst_res[0]

    def department_code(self):
        name_column = self.search_column_using_re(r'\s*код\s*подразделения\s*$')
        text = self.check_list_search(name_column)
        shablons = r'\d{3}-\d{3}'
        if not text:
            return
        for i in text:
            res = re.findall(shablons, i, re.I | re.DOTALL)
            if res:
                self.data['код подразделения'] = res[0]
                break

    def registered_address(self):
        text = self.check_list_search('адрес по прописки')
        lst_res = []
        for i in text:
            res = re.findall(r'.{20}.*',i)
            if len(res) == 0:
                continue
            lst_res.append(i)
        if lst_res:
            self.data['адрес прописки'] = lst_res[0]

    def index_parse(self):
        text = self.check_list_search('индекс')
        lst_res = []
        for i in text:
            res = re.findall(r'\s*\d{3}[-\s*]\d{3}\s*$', i)
            if len(res) == 0:
                continue
            lst_res.append(i)
        if lst_res:
            self.data['индекс'] = lst_res[0]

    def phone_number(self):
        name_column = self.search_column_using_re('\s*телефон\s*$')
        text = self.check_list_search(name_column)
        lst_res = []
        for i in text:
            res = re.findall(r'^\+7.{8}.*$|^8.{8}.*$', i)
            if len(res) == 0:
                continue
            lst_res.append(i)
        if lst_res:
            self.data['телефон'] = lst_res[0]
        else:
            print('номер телефона не найден')

    def kad_num_work(self):
        text = self.check_text_search('кадастровый номер зу/окс')
        shablons = r'\d+\s*\d+\s*\d+\s*\d+'
        res = re.findall(shablons,text)
        if res:
            self.data['кадастровый номер работ'] = res[0].replace(' ',':')


    def working_type(self):
        shablons_name_column = [r'\s*выполнить\s*раб.+\s*']
        lower_lst = self.apply_lower_methon_on_list()
        res_lst = []
        for i in lower_lst:
            for j in shablons_name_column:
                res = re.findall(j,i)
                if res:
                    res_lst += res
                    break
        if not res_lst:
            print('колонка с типом работ не найдена')
            return
        lower_lst = self.check_list_search(res_lst[0])
        shablons_mezh_plan = [r'\s*изгот.+\s*меж.*\s*плана\s*$']
        shablons_tech_work = [r'\s*изгот.+\s*техн.+\s*плана\s*$']
        res_lst = []
        for i in lower_lst:
            for j in shablons_mezh_plan:
                res = re.findall(j,i)
                if res != []:
                    res_lst += res
                    self.data['вид работ'] = 'кадастровые работы; инженерно геодезические работы'
                    return

        for i in lower_lst:
            for j in shablons_tech_work:
                res = re.findall(j, i)
                if res != []:
                    res_lst += res

                    self.data['вид работ'] = 'кадастровые работы'
                    return


    def date_contract(self):
        # с помощью регулярного выражения находим название колонки ибо она может иногда иметь лишние
        # данные
        name_column = self.search_column_using_re(r'\s*дата\s*$')
        text = self.check_list_search(name_column)
        lst_val = []
        print(text)
        for i in text:
            res = re.findall(r'\d+-\d+-\d+.*', i)
            if len(res) == 0:
                continue
            lst_val.append(i)
        if lst_val:
            self.data['дата договора'] = lst_val[0]

    def contract_number(self):
        shablons_name_column = [r'^\s*номер\s*$']
        name = None
        shablons_data = [r'\d+\s*\d+']
        for i in self.text_data:
            for j in shablons_name_column:
                res = re.findall(j,i)
                if res:
                    name = res[0]
        if not name:
            return
        text = ' '.join(self.check_list_search(name))
        for j in shablons_data:
            res = re.findall(j, text)
            if res:
                self.data['номер договора'] = res[0]
                break





    def cost_contract_ooo(self):
        text = self.check_list_search('ооо')
        lst_val = []
        for i in text:
            res = re.findall(r'\d*$', i)
            if len(res) == 0:
                continue
            lst_val.append(i)
        if lst_val:
            self.data['стоимость работ ООО'] = lst_val[0]



    def cost_contract_ip(self):
        text = self.check_list_search('ип')
        lst_val = []
        for i in text:
            res = re.findall(r'\d*$', i)
            if len(res) == 0:
                continue
            lst_val.append(i)
        if lst_val:
            self.data['стоимость работ ИП'] = lst_val[0]


    def check_text_search(self,name_column:str):
        """Метод который ищет все значения начиная от текущего столбца по которому ищется значение предваритель делая каждую букву в тексте
        маленькой"""
        lower_lst = self.apply_lower_method_on_text()
        try:
            text = ''.join(lower_lst[lower_lst.index(name_column):])
        except:
            print(f'ОШИБКА: ячейка с именем {name_column} не найдена')
            return self.text_data_text
        return text

    def check_list_search(self,name_column):
        '''метод который возыращает список ячеек начиная от текущего столбца по которому ведется поиск
            а также делает все буквы маленькими'''
        lower_lst = self.apply_lower_methon_on_list()
        try:
            text = lower_lst[lower_lst.index(name_column)+1:]
        except ValueError:
            print(f'ОШИБКА: ячейка с именем {name_column} не найдена')
            return self.text_data
        return text

    def search_column_using_re(self,template):
        res_values = []
        for i in self.apply_lower_methon_on_list():
            res = re.findall(template,i)
            if res != []:
                res_values.append(i)
        if len(res_values) == 0:
            print('ошибка: такого имени колонки нет')
            return
        return res_values[0]


    def delete_extracted_data(self,column,value):
        self.text_data.remove(column)
        self.text_data.remove(value)

    def apply_lower_method_on_text(self):
        res_str = ''''''
        for i in self.text_data:
            res_str += i.lower() + ' '
        return res_str

    def apply_lower_methon_on_list(self):
        lst = []
        for i in self.text_data:
            lst.append(i.lower())
        return lst

    def delete_concrete_values(self,value):
        '''функция которая будет удалять определенные значения в тексте, чтобы сделать его меньше и более распутаннее'''
        self.text_data_text.replace(value,'')
        self.text_data.remove(value)



    def searching_func_base(self):
        # шадлоны для поиска колонки
        shablons_name_column = []
        # шаблоны для поиска значения
        shablons_name_value = []




test = ExcelParsing(r'C:\Users\Slava-Stat\Desktop\Проекты_Python\Excel_parsing\test_data\юр лица 2023\2023-0056 на ип Войлочная Фабрика 5000\!!!New-ЮЛ ООО ДОГОВОР 2022-003 (Автосохраненный).xlsm')
list_values = test.main_call()
write_in_file.ExcelWrite().write(list_values)